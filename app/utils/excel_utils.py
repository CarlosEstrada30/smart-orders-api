import pandas as pd
import io
from typing import List, Dict, Any
from fastapi import UploadFile, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelProcessor:
    """Utility class for processing Excel files"""

    @staticmethod
    def validate_excel_file(file: UploadFile) -> None:
        """Validate if the uploaded file is an Excel file"""
        allowed_extensions = ['.xlsx', '.xls']
        if not any(file.filename.endswith(ext) for ext in allowed_extensions):
            raise HTTPException(
                status_code=400,
                detail="File must be an Excel file (.xlsx or .xls)"
            )

    @staticmethod
    async def read_excel_to_dataframe(file: UploadFile, sheet_name: str = None) -> pd.DataFrame:
        """Read Excel file and return as pandas DataFrame"""
        try:
            content = await file.read()
            df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name)
            return df
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Error reading Excel file: {str(e)}"
            )

    @staticmethod
    def validate_required_columns(df: pd.DataFrame, required_columns: List[str]) -> List[str]:
        """Validate that required columns exist in the DataFrame"""
        missing_columns = []
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        return missing_columns

    @staticmethod
    def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """Clean the DataFrame by removing empty rows and handling NaN values"""
        # Remove completely empty rows
        df = df.dropna(how='all')

        # Fill NaN values with appropriate defaults
        df = df.fillna('')

        # Strip whitespace from string columns
        string_columns = df.select_dtypes(include=['object']).columns
        for col in string_columns:
            df[col] = df[col].astype(str).str.strip()
            # Replace 'nan' string with empty string
            df[col] = df[col].replace('nan', '')

        return df


class ExcelGenerator:
    """Utility class for generating Excel files"""

    @staticmethod
    def create_clients_template() -> bytes:
        """Create Excel template for clients bulk upload"""
        # Define template data using Spanish column names
        template_data = {
            'nombre': ['Ejemplo Cliente 1', 'Ejemplo Cliente 2'],
            'email': ['cliente1@ejemplo.com', 'cliente2@ejemplo.com'],
            'teléfono': ['12345678', '87654321'],
            'nit': ['123456789', '987654321'],
            'dirección': ['Dirección ejemplo 1', 'Dirección ejemplo 2'],
            'activo': [True, True]
        }

        df = pd.DataFrame(template_data)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Add headers with styling - using Spanish column names for better UX
        headers = ['nombre', 'email', 'teléfono', 'nit', 'dirección', 'activo']
        header_descriptions = [
            'Nombre (Requerido)',
            'Email (Opcional)',
            'Teléfono (Opcional)',
            'NIT (Opcional)',
            'Dirección (Opcional)',
            'Activo (true/false)'
        ]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        description_font = Font(italic=True, size=9)
        description_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        # Add headers (row 1) with just the column names
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add descriptions (row 2) for clarity
        for idx, description in enumerate(header_descriptions, 1):
            cell = ws.cell(row=2, column=idx, value=description)
            cell.font = description_font
            cell.fill = description_fill
            cell.alignment = header_alignment

        # Add example data starting from row 3
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instrucciones")
        instructions = [
            "INSTRUCCIONES PARA CARGA MASIVA DE CLIENTES",
            "",
            "1. Complete la información en la hoja 'Clientes'",
            "2. Campos requeridos:",
            "   - nombre: Nombre del cliente (obligatorio)",
            "",
            "3. Campos opcionales:",
            "   - email: Correo electrónico del cliente",
            "   - teléfono: Número de teléfono",
            "   - nit: Número de identificación tributaria",
            "   - dirección: Dirección del cliente",
            "   - activo: true para activo, false para inactivo (por defecto: true)",
            "",
            "4. Guarde el archivo y súbalo usando el endpoint de carga masiva",
            "",
            "NOTAS IMPORTANTES:",
            "- Puede usar nombres de columnas en español o inglés",
            "- Columnas aceptadas para nombre: nombre, name, Name, Nombre, NOMBRE",
            "- Columnas aceptadas para teléfono: teléfono, telefono, phone, Phone",
            "- Columnas aceptadas para dirección: dirección, direccion, address, Address",
            "- Columnas aceptadas para activo: activo, is_active, active, Active",
            "- Puede agregar tantas filas como necesite",
            "- Los emails deben ser válidos si se proporcionan",
            "- Los valores de activo deben ser 'true' o 'false'"
        ]

        for idx, instruction in enumerate(instructions, 1):
            instructions_ws.cell(row=idx, column=1, value=instruction)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def create_products_template() -> bytes:
        """Create Excel template for products bulk upload"""
        # Define template data using Spanish column names for consistency
        template_data = {
            'nombre': ['Producto Ejemplo 1', 'Producto Ejemplo 2'],
            'descripcion': ['Descripción del producto 1', 'Descripción del producto 2'],
            'precio': [10.50, 25.00],
            'stock': [0, 0],
            'sku': ['', ''],  # Empty SKU to show it's optional
            'activo': [True, True]
        }

        df = pd.DataFrame(template_data)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Add headers with styling - using Spanish column names for consistency
        headers = ['nombre', 'descripcion', 'precio', 'stock', 'sku', 'activo']
        header_descriptions = [
            'Nombre (Requerido)',
            'Descripción (Opcional)',
            'Precio (Requerido)',
            'Stock (Opcional, por defecto: 0)',
            'SKU (Opcional, se genera automáticamente)',
            'Activo (true/false)'
        ]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        description_font = Font(italic=True, size=9)
        description_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        # Add headers (row 1) with just the column names
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add descriptions (row 2) for clarity
        for idx, description in enumerate(header_descriptions, 1):
            cell = ws.cell(row=2, column=idx, value=description)
            cell.font = description_font
            cell.fill = description_fill
            cell.alignment = header_alignment

        # Add example data starting from row 3
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instrucciones")
        instructions = [
            "INSTRUCCIONES PARA CARGA MASIVA DE PRODUCTOS",
            "",
            "1. Complete la información en la hoja 'Productos'",
            "2. Campos requeridos:",
            "   - nombre: Nombre del producto (obligatorio)",
            "   - precio: Precio del producto (obligatorio, debe ser un número)",
            "",
            "3. Campos opcionales:",
            "   - descripcion: Descripción del producto",
            "   - stock: Cantidad en inventario (por defecto: 0)",
            "   - sku: Código SKU (opcional, se genera automáticamente si no se especifica)",
            "   - activo: true para activo, false para inactivo (por defecto: true)",
            "",
            "4. Guarde el archivo y súbalo usando el endpoint de carga masiva",
            "",
            "NOTAS IMPORTANTES:",
            "- Puede usar nombres de columnas en español o inglés",
            "- Columnas aceptadas para nombre: nombre, name, Name, Nombre, NOMBRE",
            "- Columnas aceptadas para precio: precio, price, Price, Precio, PRECIO",
            "- Columnas aceptadas para descripción: descripcion, description, Description",
            "- Columnas aceptadas para activo: activo, is_active, active, Active",
            "- Puede agregar tantas filas como necesite",
            "- Los precios deben ser números válidos mayores a 0",
            "- Si no especifica SKU, se generará automáticamente",
            "- Los valores de stock deben ser números enteros (por defecto: 0)",
            "- Los valores de activo deben ser 'true' o 'false'"
        ]

        for idx, instruction in enumerate(instructions, 1):
            instructions_ws.cell(row=idx, column=1, value=instruction)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_clients_data(clients: List[Dict[str, Any]]) -> bytes:
        """Export clients data to Excel file"""
        if not clients:
            # If no clients, create empty template
            return ExcelGenerator.create_clients_template()

        # Convert clients data to the template format
        df = ExcelGenerator._prepare_clients_dataframe(clients)

        # Create workbook and add data
        wb = ExcelGenerator._create_clients_workbook(df, clients)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _prepare_clients_dataframe(clients: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare clients data in template format"""
        template_data = {
            'nombre': [],
            'email': [],
            'teléfono': [],
            'nit': [],
            'dirección': [],
            'activo': []
        }

        for client in clients:
            template_data['nombre'].append(client.get('name', ''))
            template_data['email'].append(client.get('email', '') if client.get('email') else '')
            template_data['teléfono'].append(client.get('phone', '') if client.get('phone') else '')
            template_data['nit'].append(client.get('nit', '') if client.get('nit') else '')
            template_data['dirección'].append(client.get('address', '') if client.get('address') else '')
            template_data['activo'].append(client.get('is_active', True))

        return pd.DataFrame(template_data)

    @staticmethod
    def _create_clients_workbook(df: pd.DataFrame, clients: List[Dict[str, Any]]) -> Workbook:
        """Create workbook with clients data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Add headers and data
        ExcelGenerator._add_clients_headers(ws)
        ExcelGenerator._add_dataframe_data(ws, df)
        ExcelGenerator._adjust_column_widths(ws)
        ExcelGenerator._add_clients_instructions(wb, clients)

        return wb

    @staticmethod
    def _add_clients_headers(ws):
        """Add headers with styling for clients sheet"""
        headers = ['nombre', 'email', 'teléfono', 'nit', 'dirección', 'activo']
        header_descriptions = [
            'Nombre (Requerido)',
            'Email (Opcional)',
            'Teléfono (Opcional)',
            'NIT (Opcional)',
            'Dirección (Opcional)',
            'Activo (true/false)'
        ]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        description_font = Font(italic=True, size=9)
        description_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        # Add headers (row 1) with just the column names
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add descriptions (row 2) for clarity
        for idx, description in enumerate(header_descriptions, 1):
            cell = ws.cell(row=2, column=idx, value=description)
            cell.font = description_font
            cell.fill = description_fill
            cell.alignment = header_alignment

    @staticmethod
    def _add_dataframe_data(ws, df: pd.DataFrame):
        """Add dataframe data to worksheet"""
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

    @staticmethod
    def _adjust_column_widths(ws):
        """Adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _add_clients_instructions(wb: Workbook, clients: List[Dict[str, Any]]):
        """Add instructions sheet for clients"""
        instructions_ws = wb.create_sheet("Instrucciones")
        instructions = [
            "EXPORTACIÓN DE CLIENTES",
            "",
            f"Archivo generado con {len(clients)} clientes",
            "Este archivo puede ser editado y re-importado usando el endpoint de carga masiva",
            "",
            "ESTRUCTURA DEL ARCHIVO:",
            "- Columna 'nombre': Nombre del cliente (requerido para importar)",
            "- Columna 'email': Email del cliente (opcional)",
            "- Columna 'teléfono': Teléfono del cliente (opcional)",
            "- Columna 'nit': NIT del cliente (opcional)",
            "- Columna 'dirección': Dirección del cliente (opcional)",
            "- Columna 'activo': Estado del cliente (true/false)",
            "",
            "PARA RE-IMPORTAR:",
            "1. Edite los datos según necesite",
            "2. Use POST /api/v1/clients/bulk-upload",
            "3. El sistema detectará automáticamente los nombres de columnas"
        ]

        for idx, instruction in enumerate(instructions, 1):
            instructions_ws.cell(row=idx, column=1, value=instruction)

    @staticmethod
    def export_products_data(products: List[Dict[str, Any]]) -> bytes:
        """Export products data to Excel file"""
        if not products:
            # If no products, create empty template
            return ExcelGenerator.create_products_template()

        # Convert products data to the template format
        df = ExcelGenerator._prepare_products_dataframe(products)

        # Create workbook and add data
        wb = ExcelGenerator._create_products_workbook(df, products)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _prepare_products_dataframe(products: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare products data in template format"""
        template_data = {
            'nombre': [],
            'descripcion': [],
            'precio': [],
            'stock': [],
            'sku': [],
            'activo': []
        }

        for product in products:
            template_data['nombre'].append(product.get('name', ''))
            template_data['descripcion'].append(product.get('description', '') if product.get('description') else '')
            template_data['precio'].append(product.get('price', 0))
            template_data['stock'].append(product.get('stock', 0))
            template_data['sku'].append(product.get('sku', ''))
            template_data['activo'].append(product.get('is_active', True))

        return pd.DataFrame(template_data)

    @staticmethod
    def _create_products_workbook(df: pd.DataFrame, products: List[Dict[str, Any]]) -> Workbook:
        """Create workbook with products data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Add headers and data
        ExcelGenerator._add_products_headers(ws)
        ExcelGenerator._add_dataframe_data(ws, df)
        ExcelGenerator._adjust_column_widths(ws)
        ExcelGenerator._add_products_instructions(wb, products)

        return wb

    @staticmethod
    def _add_products_headers(ws):
        """Add headers with styling for products sheet"""
        headers = ['nombre', 'descripcion', 'precio', 'stock', 'sku', 'activo']
        header_descriptions = [
            'Nombre (Requerido)',
            'Descripción (Opcional)',
            'Precio (Requerido)',
            'Stock (Opcional, por defecto: 0)',
            'SKU (Opcional, se genera automáticamente)',
            'Activo (true/false)'
        ]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        description_font = Font(italic=True, size=9)
        description_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        # Add headers (row 1) with just the column names
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add descriptions (row 2) for clarity
        for idx, description in enumerate(header_descriptions, 1):
            cell = ws.cell(row=2, column=idx, value=description)
            cell.font = description_font
            cell.fill = description_fill
            cell.alignment = header_alignment

    @staticmethod
    def _add_products_instructions(wb: Workbook, products: List[Dict[str, Any]]):
        """Add instructions sheet for products"""
        instructions_ws = wb.create_sheet("Instrucciones")
        instructions = [
            "EXPORTACIÓN DE PRODUCTOS",
            "",
            f"Archivo generado con {len(products)} productos",
            "Este archivo puede ser editado y re-importado usando el endpoint de carga masiva",
            "",
            "ESTRUCTURA DEL ARCHIVO:",
            "- Columna 'nombre': Nombre del producto (requerido para importar)",
            "- Columna 'descripcion': Descripción del producto (opcional)",
            "- Columna 'precio': Precio del producto (requerido para importar)",
            "- Columna 'stock': Stock del producto (opcional, por defecto: 0)",
            "- Columna 'sku': SKU del producto (opcional, se genera automáticamente si no se especifica)",
            "- Columna 'activo': Estado del producto (true/false)",
            "",
            "PARA RE-IMPORTAR:",
            "1. Edite los datos según necesite",
            "2. IMPORTANTE: Solo nombre y precio son requeridos",
            "3. Use POST /api/v1/products/bulk-upload",
            "4. El sistema detectará automáticamente los nombres de columnas"
        ]

        for idx, instruction in enumerate(instructions, 1):
            instructions_ws.cell(row=idx, column=1, value=instruction)
